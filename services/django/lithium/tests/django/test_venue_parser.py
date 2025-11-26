from tempfile import NamedTemporaryFile

from django.test import SimpleTestCase
from openpyxl import Workbook
from openpyxl.styles import Font

from timetabling_system.utils.venue_parser import parse_venue_file


class TestVenueParser(SimpleTestCase):
    def test_red_room_marks_accessible_true(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "Monday")
        ws.cell(2, 1, "2025-07-28")

        red_room = ws.cell(3, 1, "Accessible Room")
        red_room.font = Font(color="FF0000")
        ws.cell(4, 1, "Default Room")

        with NamedTemporaryFile(suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            result = parse_venue_file(tmp.name)

        rooms = result["days"][0]["rooms"]
        self.assertEqual(rooms[0]["name"], "Accessible Room")
        self.assertTrue(rooms[0]["accessible"])
        self.assertEqual(rooms[1]["name"], "Default Room")
        self.assertTrue(rooms[1]["accessible"])
