from openpyxl import load_workbook

def parse_venue_file(file):
    print("Parsing venue file...")
    wb = load_workbook(file)
    ws = wb.active

    results = []

    # Read column pairs: (header row, date row, data rows)
    header_row = 1
    date_row = 2
    first_data_row = 3

    for col in range(1, ws.max_column + 1):
        day_cell = ws.cell(header_row, col)
        date_cell = ws.cell(date_row, col)

        day_text = str(day_cell.value).strip() if day_cell.value else None
        date_text = str(date_cell.value).strip() if date_cell.value else None

        # Skip empty columns
        if not day_text:
            continue

        rooms = []

        for row in range(first_data_row, ws.max_row + 1):
            cell = ws.cell(row, col)
            value = cell.value

            if not value:
                continue

            # Detect red font (non-accessible)
            font_color = cell.font.color
            is_red = (
                font_color
                and font_color.type == "rgb"
                and font_color.rgb
                and font_color.rgb.upper().startswith("FF0000")
            )

            rooms.append({
                "name": str(value).strip(),
                "accessible": not is_red
            })

        results.append({
            "day": day_text,
            "date": date_text,
            "rooms": rooms
        })

    return {
        "status": "ok",
        "type": "Venue",
        "days": results
    }
